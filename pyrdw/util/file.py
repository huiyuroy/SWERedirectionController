import os
import pickle
import xml.etree.ElementTree as ET
from tkinter import filedialog

import numpy as np
import openpyxl
import ujson
from PIL import Image

import pyrdw.lib.math.geometry as geo

from pyrdw.core.space.grid import Tiling
from pyrdw.core.space.scene import DiscreteScene


def load_bin(file_path=None):
    try:
        if file_path is not None:
            with open(file_path, mode='rb') as f:
                return pickle.load(f)
    except Exception as e:
        print(e)
        return None


def save_bin(data, file_path=None):
    dir_path, name = os.path.split(file_path)
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    with open(file_path, mode='wb') as f:
        pickle.dump(data, f)


def load_json(file_path=None):
    try:
        if file_path is not None:
            with open(file_path, mode='r') as f:
                return ujson.load(f)
    except Exception as e:
        return None


def save_json(data, file_path=None):
    dir_path, name = os.path.split(file_path)
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    with open(file_path, mode='w') as f:
        ujson.dump(data, f, ensure_ascii=True, indent=2)


def save_grid_image(data, file_path):
    tiling_weight_image = Image.fromarray(data).convert('L')
    tiling_weight_image.save(file_path)


class JsonConverter:

    def __init__(self):
        self.data = None
        self.file_path = None
        self.root_path = None
        self.f_name = None

    def set_tar_json_path(self, file_path):
        self.file_path = file_path

    def gen_tar_json(self):
        if not os.path.isdir(self.file_path):
            os.mkdir(self.file_path)
        with open(self.file_path, mode='w') as f:
            ujson.dump(self.data, f)

    def converter_json(self, file_path=None):
        if file_path is not None:
            self.file_path = file_path
        with open(self.file_path, mode='r') as f:
            j_result = ujson.load(f)
            return j_result

    def save_json(self, data, file_path=None):
        self.file_path = file_path
        with open(self.file_path, 'w') as f:
            ujson.dump(data, f, ensure_ascii=True, indent=2)


class ExcelWriter:

    def __init__(self):
        self.root_path = None
        self.file_name = None
        self.tar_path = None
        self.sheet_name = None
        self.work_book = None
        self.sheet = None

    def open_excel(self, f_path, f_name):
        if not os.path.exists(f_path):
            os.makedirs(f_path)
        f_path = f_path + '\\' + f_name + '.xlsx'
        self.tar_path = f_path
        self.work_book = openpyxl.load_workbook(f_path) if os.path.exists(f_path) else openpyxl.Workbook()

    def update_sheet(self, sheet_name, titles):
        sheet_names = self.work_book.sheetnames
        if sheet_name in sheet_names:
            self.work_book.remove_sheet(self.work_book.get_sheet_by_name(sheet_name))
        self.sheet = self.work_book.create_sheet(sheet_name)
        self.sheet.append(titles)



    def write_excel(self, *args):
        # self.sheet.append(data)
        self.sheet.append(args)

    def save_excel(self):
        self.work_book.save(self.tar_path)
